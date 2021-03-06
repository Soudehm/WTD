import openpyxl,pprint
import glob
import os
from openpyxl.cell import get_column_letter, column_index_from_string
from os import listdir
from os.path import isfile, join
from os import walk

os.chdir('/Users/Soudehm/Documents/July4thWTW/WBS')
directory = "/Users/Soudehm/Documents/July4thWTW/DataBaseOfMyCloset"
wb = openpyxl.load_workbook('WTW_DB.xlsx')

#------------------------------------------------------------------------------
# this code was modified accordingly http: //stackoverflow.com/questions/3207219/how-to-list-all-files-of-a-directory-in-python

def get_filepaths(directory):
    
    #    """
    #    This function will generate the file names in a directory
    #    tree by walking the tree either top-down or bottom-up. For each
    #    directory in the tree rooted at directory top (including top itself),
    #    it yields a 3-tuple (dirpath, dirnames, filenames).
    #    """
    file_paths = []  # List which will store all of the full filepaths.
    
    # Walk the tree.
    for root, directories, files in os.walk(directory):
        for filename in files:
            # Join the two strings in order to form the full filepath.
            filepath = os.path.join(root, filename)
            file_paths.append(filepath)  # Add it to the list.

    return file_paths  # Self-explanatory.
#------------------------------------------------------------------------------
# Run the above function and store its results in a variable.
full_file_paths = get_filepaths(directory)

sheet_data = wb.active
sheet_data.title = 'WTW_DB_data'
sheet_data.sheet_properties.tabColor = "1072BA"

for index_files in range(len(full_file_paths)):
    sheet_data['A' + str(index_files + 1)].value = full_file_paths [index_files]

#------------------------------------------------------------------------------
#------------------------------------------------------------------------------
# this code was modified accordingly http: //stackoverflow.com/questions/3207219/how-to-list-all-files-of-a-directory-in-python

def get_filenames(directory):
    
    #    """
    #    This function will generate the file names in a directory
    #    tree by walking the tree either top-down or bottom-up. For each
    #    directory in the tree rooted at directory top (including top itself),
    #    it yields a 3-tuple (dirpath, dirnames, filenames).
    #    """
    file_names = []  # List which will store all of the full filepaths.
    
    # Walk the tree.
    for root, directories, files in os.walk(directory):
        file_names.append(files)

    return file_names  # Self-explanatory.
#------------------------------------------------------------------------------
# Run the above function and store its results in a variable.
full_file_names = get_filenames(directory)
#print(full_file_names)

workplace = wb.create_sheet() # insert a sheet at the end
workplace.title = 'WTW_DB'
workplace.sheet_properties.tabColor = "BA72BA"

for index_files in range(len(full_file_names)):
    workplace['A' + str(index_files + 1)].value = full_file_names[index_files]

wb.save('WTW_DB.xlsx')